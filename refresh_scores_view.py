"""
PPB26 Refresh Script
====================
Run this script whenever you have imported new responses into Import_Responses.

It rebuilds:
  - Scores_View   (the pivot layout: fields as rows, reviewers as columns)
  - Graph Summary (ranked bar chart data table — chart auto-updates)

Everything else updates automatically in Excel via live formulas:
  - Scores            (uses ROW()-based INDEX/MATCH formulas)
  - Summary           (uses COUNTIF / AVERAGEIFS on Scores)
  - Reviewer_Summary  (uses COUNTIF / AVERAGEIFS on Scores and Import_Responses)

Usage:
  pip3 install openpyxl
  python3 refresh_scores_view.py

Place this script in the same folder as the workbook, or edit WORKBOOK_PATH below.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

WORKBOOK_PATH = 'PPB26_RFP_Scoring_AutoRank_Template.xlsx'

# ── Styles ─────────────────────────────────────────────────────────────────────
thin  = Side(style='thin')
bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)

PROPOSAL_FILL  = PatternFill('solid', start_color='2E4057')
PROPOSAL_FONT  = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
PROPOSAL_ALIGN = Alignment(horizontal='left', vertical='center')
REVIEWER_FILL  = PatternFill('solid', start_color='4472C4')
REVIEWER_FONT  = Font(name='Calibri', bold=True, size=10, color='FFFFFF')
REVIEWER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
FIELD_FILL     = PatternFill('solid', start_color='D9E1F2')
FIELD_FONT     = Font(name='Calibri', bold=True, size=10)
FIELD_ALIGN    = Alignment(horizontal='left', vertical='center')
DATA_FONT      = Font(name='Calibri', size=10)
HEADER_FONT    = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
HEADER_FILL    = PatternFill('solid', start_color='4472C4')
HEADER_ALIGN   = Alignment(horizontal='center', vertical='center', wrap_text=True)
TITLE_FONT     = Font(name='Calibri', bold=True, size=14, color='FFFFFF')
TITLE_FILL     = PatternFill('solid', start_color='2E4057')
ALT_FILL       = PatternFill('solid', start_color='EEF2FF')
CTR            = Alignment(horizontal='center', vertical='center')
LEFT           = Alignment(horizontal='left', vertical='center')
TXT_ALIGN      = Alignment(horizontal='left', vertical='top', wrap_text=True)

FIELD_NAMES = [
    'Recused (1/0)', 'Q1 (0-4)', 'Q2 (0-4)', 'Q3 (0-4)', 'Q4 (0-4)',
    'Q5 (0-4)', 'Total Points', '# Questions Scored', 'Average Score (0-4)',
    '% of Max', 'Overall Strengths', 'Key Concerns', 'Overall Recommendation'
]


def read_scores(wb):
    """Read calculated values from the Scores sheet.
    Includes comment-only rows (no Q scores) as long as proposal + reviewer are present.
    """
    ws = wb['Scores']
    rows = []
    empty_streak = 0
    for r in range(2, 9999):
        proposal = ws.cell(r, 2).value
        reviewer = ws.cell(r, 3).value
        if not proposal or not reviewer:
            empty_streak += 1
            if empty_streak >= 10:
                break
            continue
        empty_streak = 0
        rows.append({
            'proposal': str(proposal).replace('\xa0', ' ').strip(),
            'reviewer': str(reviewer).strip(),
            'vals': [ws.cell(r, c).value for c in range(4, 17)]
        })
    return rows


def build_scores_view(wb, rows):
    """Rebuild the Scores_View pivot sheet."""
    proposals, seen = [], set()
    for r in rows:
        p = r['proposal']
        if p not in seen:
            proposals.append(p)
            seen.add(p)

    reviewers = sorted(set(r['reviewer'] for r in rows))
    n_rev = len(reviewers)
    last_col = 1 + n_rev

    data = {p: {rev: [None]*13 for rev in reviewers} for p in proposals}
    for r in rows:
        data[r['proposal']][r['reviewer']] = r['vals']

    if 'Scores_View' in wb.sheetnames:
        del wb['Scores_View']
    ws = wb.create_sheet('Scores_View', 5)

    ws.column_dimensions['A'].width = 22
    for ri in range(n_rev):
        ws.column_dimensions[get_column_letter(2 + ri)].width = 24

    current_row = 1
    for proposal in proposals:
        # Reviewer header row
        ws.cell(current_row, 1).fill = REVIEWER_FILL
        ws.cell(current_row, 1).border = bdr
        for ri, rev in enumerate(reviewers):
            c = ws.cell(current_row, 2 + ri)
            c.value, c.font, c.fill, c.alignment, c.border = (
                rev, REVIEWER_FONT, REVIEWER_FILL, REVIEWER_ALIGN, bdr)
        ws.row_dimensions[current_row].height = 40
        current_row += 1

        # Proposal title row (merged)
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=last_col)
        c = ws.cell(current_row, 1)
        c.value, c.font, c.fill, c.alignment, c.border = (
            proposal, PROPOSAL_FONT, PROPOSAL_FILL, PROPOSAL_ALIGN, bdr)
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # One row per field
        for fi, field in enumerate(FIELD_NAMES):
            is_text = fi >= 10
            c = ws.cell(current_row, 1)
            c.value, c.font, c.fill, c.alignment, c.border = (
                field, FIELD_FONT, FIELD_FILL, FIELD_ALIGN, bdr)
            for ri, rev in enumerate(reviewers):
                val  = data[proposal][rev][fi]
                cell = ws.cell(current_row, 2 + ri)
                cell.value, cell.font, cell.border = val, DATA_FONT, bdr
                cell.alignment = TXT_ALIGN if is_text else CTR
                if fi == 9:   cell.number_format = '0.0%'
                elif fi == 8: cell.number_format = '0.00'
            ws.row_dimensions[current_row].height = 60 if is_text else 16
            current_row += 1

        current_row += 1  # gap row between proposals

    ws.freeze_panes = 'B1'
    print(f"  Scores_View: {len(proposals)} proposals × {n_rev} reviewers, {current_row} rows")


def build_graph_summary(wb, rows):
    """Rebuild the Graph Summary data table and chart, calculated directly from scored rows."""
    ws_gs = wb['Graph Summary']

    # Clear existing charts and merges
    ws_gs._charts = []
    for m in list(ws_gs.merged_cells.ranges):
        ws_gs.unmerge_cells(str(m))

    # Title
    ws_gs.merge_cells('A1:C1')
    c = ws_gs.cell(1, 1)
    c.value = 'PPB26 Proposal Ranking — Average Score (0-4)'
    c.font, c.fill, c.alignment = TITLE_FONT, TITLE_FILL, Alignment(horizontal='center', vertical='center')
    ws_gs.row_dimensions[1].height = 30

    # Headers
    for ci, (h, w) in enumerate(zip(['Rank', 'Proposal', 'Avg Score (0-4)'], [7, 52, 16]), 1):
        c = ws_gs.cell(2, ci)
        c.value, c.font, c.fill, c.alignment, c.border = h, HEADER_FONT, HEADER_FILL, HEADER_ALIGN, bdr
        ws_gs.column_dimensions[get_column_letter(ci)].width = w
    ws_gs.row_dimensions[2].height = 30

    # Calculate avg score per proposal from rows data
    # vals index: 0=Recused, 1-5=Q1-Q5, 6=Total, 7=#Qs, 8=Avg(0-4), 9=%Max ...
    # Only include non-recused rows that have an actual numeric average (index 8)
    from collections import defaultdict
    proposal_scores = defaultdict(list)
    for r in rows:
        recused = r['vals'][0]
        avg_val = r['vals'][8]
        if recused != 1 and avg_val is not None:
            try:
                proposal_scores[r['proposal']].append(float(avg_val))
            except (TypeError, ValueError):
                pass  # skip formula strings or blank cells

    # Build ranked list — only proposals with at least one valid score
    proposals_sum = []
    for proposal, scores in proposal_scores.items():
        if scores:
            avg = sum(scores) / len(scores)
            proposals_sum.append({'name': proposal, 'avg': avg, 'n': len(scores)})

    # Rank by avg score descending, then n_responses descending
    proposals_sum.sort(key=lambda x: (-x['avg'], -x['n']))
    # Assign ranks (handle ties)
    rank = 1
    for i, p in enumerate(proposals_sum):
        if i > 0 and p['avg'] == proposals_sum[i-1]['avg']:
            p['rank'] = proposals_sum[i-1]['rank']
        else:
            p['rank'] = rank
        rank = i + 2

    n = len(proposals_sum)

    # Write data rows reversed (worst→best so rank 1 appears at top of horizontal bar)
    for i, p in enumerate(reversed(proposals_sum)):
        row  = 3 + i
        fill = ALT_FILL if i % 2 == 0 else PatternFill()
        ws_gs.row_dimensions[row].height = 16

        c1 = ws_gs.cell(row, 1)
        c1.value, c1.font, c1.fill, c1.alignment, c1.border = p['rank'], DATA_FONT, fill, CTR, bdr

        c2 = ws_gs.cell(row, 2)
        c2.value = str(p['name']).replace('\xa0', ' ').strip()[:55]
        c2.font, c2.fill, c2.alignment, c2.border = DATA_FONT, fill, LEFT, bdr

        c3 = ws_gs.cell(row, 3)
        c3.value, c3.font, c3.fill, c3.alignment, c3.border = round(p['avg'], 3), DATA_FONT, fill, CTR, bdr
        c3.number_format = '0.00'

    # Chart
    chart = BarChart()
    chart.type, chart.grouping = 'bar', 'clustered'
    chart.title = None
    chart.x_axis.title = 'Average Score (0-4)'
    chart.x_axis.scaling.min, chart.x_axis.scaling.max = 0, 4
    chart.style, chart.width, chart.height = 10, 22, 22

    data_ref = Reference(ws_gs, min_col=3, min_row=3, max_row=2 + n)
    chart.add_data(data_ref)
    cats = Reference(ws_gs, min_col=2, min_row=3, max_row=2 + n)
    chart.set_categories(cats)
    chart.series[0].title = None
    chart.series[0].graphicalProperties.solidFill = '4472C4'
    ws_gs.add_chart(chart, 'E2')

    print(f"  Graph Summary: {n} proposals ranked")


def main():
    print(f"\nLoading {WORKBOOK_PATH} ...")
    # Load with data_only to get calculated values from Scores
    wb_data = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
    rows = read_scores(wb_data)
    print(f"  Found {len(rows)} scored responses")

    # Load formula version to preserve all other formulas
    wb = openpyxl.load_workbook(WORKBOOK_PATH)

    print("\nRebuilding Scores_View ...")
    build_scores_view(wb, rows)

    print("Rebuilding Graph Summary ...")
    build_graph_summary(wb, rows)

    wb.save(WORKBOOK_PATH)
    print(f"\nDone! Saved to {WORKBOOK_PATH}\n")


if __name__ == '__main__':
    main()
